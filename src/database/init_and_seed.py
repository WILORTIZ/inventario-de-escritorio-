import sqlite3
import openpyxl
import os
import datetime

DB_PATH = r"D:\antigravity\INVENTARIO\inventario.db"
EXCEL_PATH = r"D:\antigravity\INVENTARIO\Inventario_Gio.xlsm"

def init_db():
    if os.path.exists(DB_PATH):
        try:
            os.remove(DB_PATH)
        except Exception as e:
            print("Notice on remove DB:", e)
            
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    # 1. Tabla de Items
    cur.execute('''
    CREATE TABLE IF NOT EXISTS items (
        codigo INTEGER PRIMARY KEY,
        nombre TEXT NOT NULL,
        categoria TEXT NOT NULL,
        subcategoria TEXT DEFAULT 'General',
        unidad_medida TEXT NOT NULL,
        marca TEXT DEFAULT 'Generico',
        referencia TEXT DEFAULT '-',
        ubicacion_cds TEXT DEFAULT 'A1',
        aplica_vencimiento INTEGER DEFAULT 0,
        fecha_vencimiento_default TEXT,
        stock_minimo INTEGER DEFAULT 0,
        estado TEXT DEFAULT 'Activo',
        observaciones TEXT,
        fecha_registro TEXT
    )
    ''')
    
    # 2. Tabla de Bodegas
    cur.execute('''
    CREATE TABLE IF NOT EXISTS bodegas (
        codigo TEXT PRIMARY KEY,
        nombre TEXT NOT NULL UNIQUE,
        ubicacion TEXT,
        responsable TEXT,
        estado TEXT DEFAULT 'Activa',
        observaciones TEXT
    )
    ''')
    
    # 3. Tabla de Proyectos
    cur.execute('''
    CREATE TABLE IF NOT EXISTS proyectos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL UNIQUE,
        responsable TEXT,
        estado TEXT DEFAULT 'Activo',
        observaciones TEXT
    )
    ''')
    
    # 4. Tabla de Movimientos
    cur.execute('''
    CREATE TABLE IF NOT EXISTS movimientos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        n_movimiento TEXT NOT NULL UNIQUE,
        fecha TEXT NOT NULL,
        hora TEXT NOT NULL,
        tipo_movimiento TEXT NOT NULL,
        codigo_item INTEGER NOT NULL,
        nombre_item TEXT NOT NULL,
        cantidad REAL NOT NULL,
        unidad TEXT NOT NULL,
        bodega_origen TEXT,
        bodega_destino TEXT,
        causal_condicion TEXT,
        ubicacion_cds TEXT,
        proyecto_destino TEXT,
        responsable TEXT,
        persona_recibe_devuelve TEXT,
        documento_referencia TEXT,
        observaciones TEXT,
        fecha_vencimiento_lote TEXT,
        creado_en DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (codigo_item) REFERENCES items(codigo)
    )
    ''')
    
    # 5. Tabla de Control de Vencimientos
    cur.execute('''
    CREATE TABLE IF NOT EXISTS control_vencimientos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_item INTEGER NOT NULL,
        nombre_item TEXT NOT NULL,
        bodega TEXT DEFAULT 'CDS',
        fecha_ingreso TEXT NOT NULL,
        fecha_vencimiento TEXT NOT NULL,
        cant_inicial REAL NOT NULL,
        cant_disponible REAL NOT NULL,
        estado TEXT,
        observaciones TEXT,
        n_movimiento_origen TEXT,
        FOREIGN KEY (codigo_item) REFERENCES items(codigo)
    )
    ''')
    
    # 6. Tabla de Listas y Configuracion
    cur.execute('''
    CREATE TABLE IF NOT EXISTS listas_config (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo TEXT NOT NULL,
        valor TEXT NOT NULL,
        orden INTEGER DEFAULT 0
    )
    ''')
    
    conn.commit()
    conn.close()
    print("Database schema created successfully.")

def clean_val(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, datetime.time):
        return val.strftime('%H:%M:%S')
    if isinstance(val, datetime.date):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    return s if s else None

def seed_data():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    # 1. BODEGAS
    if 'BODEGAS' in wb.sheetnames:
        sheet = wb['BODEGAS']
        for r in range(4, sheet.max_row + 1):
            cod = clean_val(sheet.cell(row=r, column=1).value)
            nom = clean_val(sheet.cell(row=r, column=2).value)
            ubic = clean_val(sheet.cell(row=r, column=3).value)
            resp = clean_val(sheet.cell(row=r, column=4).value)
            est = clean_val(sheet.cell(row=r, column=5).value) or 'Activa'
            obs = clean_val(sheet.cell(row=r, column=6).value)
            if cod and nom:
                cur.execute('''
                INSERT OR REPLACE INTO bodegas (codigo, nombre, ubicacion, responsable, estado, observaciones)
                VALUES (?, ?, ?, ?, ?, ?)
                ''', (cod, nom, ubic, resp, est, obs))
        print("Bodegas imported.")
        
    # 2. LISTAS_CONFIG y PROYECTOS
    if 'LISTAS_CONFIG' in wb.sheetnames:
        sheet = wb['LISTAS_CONFIG']
        type_cols = {
            1: 'categoria',
            2: 'unidad_medida',
            3: 'tipo_movimiento',
            7: 'ubicacion_cds',
            8: 'causal_disposicion'
        }
        
        for col_idx, tipo in type_cols.items():
            for r in range(2, sheet.max_row + 1):
                val = clean_val(sheet.cell(row=r, column=col_idx).value)
                if val:
                    cur.execute('''
                    INSERT INTO listas_config (tipo, valor, orden)
                    VALUES (?, ?, ?)
                    ''', (tipo, val, r))
                    
        # Proyectos (Col 6)
        for r in range(2, sheet.max_row + 1):
            val = clean_val(sheet.cell(row=r, column=6).value)
            if val:
                cur.execute('''
                INSERT OR IGNORE INTO proyectos (nombre, estado)
                VALUES (?, 'Activo')
                ''', (val,))
        print("Lists and Projects imported.")
        
    # 3. ITEMS
    if 'ITEMS' in wb.sheetnames:
        sheet = wb['ITEMS']
        items_count = 0
        for r in range(4, sheet.max_row + 1):
            raw_cod = sheet.cell(row=r, column=1).value
            if raw_cod is None:
                continue
            try:
                codigo = int(float(str(raw_cod).strip()))
            except:
                continue
                
            nombre = clean_val(sheet.cell(row=r, column=2).value)
            if not nombre:
                continue
                
            categoria = clean_val(sheet.cell(row=r, column=3).value) or 'Materiales'
            subcategoria = clean_val(sheet.cell(row=r, column=4).value) or 'General'
            unidad = clean_val(sheet.cell(row=r, column=5).value) or 'Unidad'
            marca = clean_val(sheet.cell(row=r, column=6).value) or 'Generico'
            referencia = clean_val(sheet.cell(row=r, column=7).value) or '-'
            ubicacion = clean_val(sheet.cell(row=r, column=8).value) or 'A1'
            f_venc = clean_val(sheet.cell(row=r, column=9).value)
            aplica_venc = 1 if f_venc and str(f_venc).upper() not in ['NO APLICA', 'NO', 'N/A', '-'] else 0
            
            raw_min = sheet.cell(row=r, column=10).value
            try:
                stock_min = int(float(str(raw_min))) if raw_min is not None else 0
            except:
                stock_min = 0
                
            estado = clean_val(sheet.cell(row=r, column=11).value) or 'Activo'
            obs = clean_val(sheet.cell(row=r, column=12).value)
            f_reg = clean_val(sheet.cell(row=r, column=13).value) or '2026-08-29'
            
            cur.execute('''
            INSERT OR REPLACE INTO items (codigo, nombre, categoria, subcategoria, unidad_medida, marca, referencia, ubicacion_cds, aplica_vencimiento, fecha_vencimiento_default, stock_minimo, estado, observaciones, fecha_registro)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (codigo, nombre, categoria, subcategoria, unidad, marca, referencia, ubicacion, aplica_venc, f_venc, stock_min, estado, obs, f_reg))
            items_count += 1
        print(f"Items imported: {items_count}")
        
    # 4. HISTORIAL_MOVIMIENTOS
    if 'HISTORIAL_MOVIMIENTOS' in wb.sheetnames:
        sheet = wb['HISTORIAL_MOVIMIENTOS']
        movs_count = 0
        for r in range(4, sheet.max_row + 1):
            n_mov = clean_val(sheet.cell(row=r, column=1).value)
            if not n_mov:
                continue
                
            f_val = clean_val(sheet.cell(row=r, column=2).value) or '2026-08-29'
            h_val = clean_val(sheet.cell(row=r, column=3).value) or '08:00:00'
            tipo_mov = clean_val(sheet.cell(row=r, column=4).value) or 'ENTRADA'
            
            raw_cod = sheet.cell(row=r, column=5).value
            try:
                cod_item = int(float(str(raw_cod).strip()))
            except:
                continue
                
            nom_item = clean_val(sheet.cell(row=r, column=6).value) or ''
            
            raw_cant = sheet.cell(row=r, column=7).value
            try:
                cant = float(str(raw_cant).strip()) if raw_cant is not None else 0.0
            except:
                cant = 0.0
                
            unidad = clean_val(sheet.cell(row=r, column=8).value) or 'Unidad'
            b_orig = clean_val(sheet.cell(row=r, column=9).value)
            b_dest = clean_val(sheet.cell(row=r, column=10).value)
            causal = clean_val(sheet.cell(row=r, column=11).value)
            ubic = clean_val(sheet.cell(row=r, column=12).value)
            proy = clean_val(sheet.cell(row=r, column=13).value)
            resp = clean_val(sheet.cell(row=r, column=14).value)
            recibe = clean_val(sheet.cell(row=r, column=15).value)
            doc_ref = clean_val(sheet.cell(row=r, column=16).value)
            obs = clean_val(sheet.cell(row=r, column=17).value)
            
            cur.execute('''
            INSERT OR REPLACE INTO movimientos (
                n_movimiento, fecha, hora, tipo_movimiento, codigo_item, nombre_item, cantidad, unidad,
                bodega_origen, bodega_destino, causal_condicion, ubicacion_cds, proyecto_destino,
                responsable, persona_recibe_devuelve, documento_referencia, observaciones
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (n_mov, f_val, h_val, tipo_mov, cod_item, nom_item, cant, unidad, b_orig, b_dest, causal, ubic, proy, resp, recibe, doc_ref, obs))
            movs_count += 1
        print(f"Movements imported: {movs_count}")

    # 5. CONTROL_VENCIMIENTOS
    if 'CONTROL_VENCIMIENTOS' in wb.sheetnames:
        sheet = wb['CONTROL_VENCIMIENTOS']
        venc_count = 0
        for r in range(4, sheet.max_row + 1):
            raw_cod = sheet.cell(row=r, column=1).value
            if not raw_cod:
                continue
            try:
                cod_item = int(float(str(raw_cod).strip()))
            except:
                continue
            nom_item = clean_val(sheet.cell(row=r, column=2).value)
            bodega = clean_val(sheet.cell(row=r, column=3).value) or 'CDS'
            f_ing = clean_val(sheet.cell(row=r, column=4).value)
            f_venc = clean_val(sheet.cell(row=r, column=5).value)
            c_ini = float(sheet.cell(row=r, column=6).value or 0)
            c_disp = float(sheet.cell(row=r, column=7).value or 0)
            estado = clean_val(sheet.cell(row=r, column=8).value)
            obs = clean_val(sheet.cell(row=r, column=9).value)
            
            cur.execute('''
            INSERT INTO control_vencimientos (
                codigo_item, nombre_item, bodega, fecha_ingreso, fecha_vencimiento,
                cant_inicial, cant_disponible, estado, observaciones
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (cod_item, nom_item, bodega, f_ing, f_venc, c_ini, c_disp, estado, obs))
            venc_count += 1
        print(f"Expiration batches imported: {venc_count}")

    conn.commit()
    conn.close()
    print("Database seeding completed successfully!")

if __name__ == '__main__':
    init_db()
    seed_data()
